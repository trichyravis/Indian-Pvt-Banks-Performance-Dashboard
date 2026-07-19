"""Indian Private Banks — Performance Analytics.

A classroom-ready Streamlit dashboard comparing five leading Indian private-sector
banks across five audited years plus the latest quarter, with a composite scorecard,
peer radar, growth analytics, asset-quality relationships, live NSE returns, auto
commentary, a ratio glossary and Excel/CSV/profile downloads.

Author: Prof. V. Ravichandran — The Mountain Path Academy.
"""
from __future__ import annotations

from io import BytesIO
from pathlib import Path

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl.styles import Font, PatternFill

# --------------------------------------------------------------------------- #
# Configuration and constants
# --------------------------------------------------------------------------- #
ROOT = Path(__file__).resolve().parent
DATA = ROOT / "data"

BANK_COLORS = {
    "ICICI Bank": "#F58220",
    "HDFC Bank": "#004C8F",
    "Axis Bank": "#97144D",
    "Kotak Mahindra Bank": "#ED1C24",
    "Yes Bank": "#005BAA",
}
TICKERS = {
    "ICICI Bank": "ICICIBANK.NS",
    "HDFC Bank": "HDFCBANK.NS",
    "Axis Bank": "AXISBANK.NS",
    "Kotak Mahindra Bank": "KOTAKBANK.NS",
    "Yes Bank": "YESBANK.NS",
}

# label -> (column, higher_is_better, value_format)
METRICS: dict[str, tuple[str, bool, str]] = {
    "Net Profit (₹ crore)": ("net_profit_cr", True, "₹{:,.0f} Cr"),
    "Net Interest Income (₹ crore)": ("nii_cr", True, "₹{:,.0f} Cr"),
    "Net Interest Margin (%)": ("nim_pct", True, "{:.2f}%"),
    "Advances (₹ crore)": ("advances_cr", True, "₹{:,.0f} Cr"),
    "Deposits (₹ crore)": ("deposits_cr", True, "₹{:,.0f} Cr"),
    "CASA Ratio (%)": ("casa_pct", True, "{:.2f}%"),
    "Gross NPA (%)": ("gnpa_pct", False, "{:.2f}%"),
    "Net NPA (%)": ("nnpa_pct", False, "{:.2f}%"),
    "Return on Assets (%)": ("roa_pct", True, "{:.2f}%"),
    "Return on Equity (%)": ("roe_pct", True, "{:.2f}%"),
    "Capital Adequacy (%)": ("car_pct", True, "{:.2f}%"),
    "Cost-to-Income (%)": ("cost_income_pct", False, "{:.2f}%"),
}

# Metrics that feed the composite scorecard: (category, column, higher_is_better)
SCORE_METRICS: list[tuple[str, str, bool]] = [
    ("Profitability", "roa_pct", True),
    ("Profitability", "roe_pct", True),
    ("Margin", "nim_pct", True),
    ("Funding", "casa_pct", True),
    ("Asset quality", "gnpa_pct", False),
    ("Asset quality", "nnpa_pct", False),
    ("Capital", "car_pct", True),
    ("Efficiency", "cost_income_pct", False),
]

# Radar dimensions: (display label, column, higher_is_better)
RADAR_METRICS: list[tuple[str, str, bool]] = [
    ("NIM", "nim_pct", True),
    ("ROA", "roa_pct", True),
    ("ROE", "roe_pct", True),
    ("CASA", "casa_pct", True),
    ("Asset quality", "gnpa_pct", False),
    ("Capital", "car_pct", True),
]

CAGR_METRICS: dict[str, str] = {
    "Net profit": "net_profit_cr",
    "Net interest income": "nii_cr",
    "Advances": "advances_cr",
    "Deposits": "deposits_cr",
}

GLOSSARY: dict[str, str] = {
    "Net Interest Income (NII)": "Interest earned on advances and investments minus interest "
        "paid on deposits and borrowings — the core earnings engine of a bank.",
    "Net Interest Margin (NIM)": "NII expressed as a percentage of average interest-earning "
        "assets. It measures how profitably a bank lends. Higher is generally better.",
    "CASA Ratio": "Share of low-cost Current and Savings Account deposits in total deposits. "
        "A higher CASA lowers the cost of funds, but must be read with overall deposit growth.",
    "Gross NPA (%)": "Gross non-performing assets as a share of gross advances — the proportion "
        "of the loan book that has stopped performing. Lower is better.",
    "Net NPA (%)": "Gross NPAs less provisions, as a share of net advances. It reflects the "
        "unprovided credit risk still on the balance sheet. Lower is better.",
    "Return on Assets (ROA)": "Net profit as a percentage of average total assets — how "
        "efficiently the bank turns its asset base into profit.",
    "Return on Equity (ROE)": "Net profit as a percentage of shareholders' equity — the return "
        "generated for owners. Read alongside leverage and capital adequacy.",
    "Capital Adequacy (CAR)": "Capital as a percentage of risk-weighted assets (Basel III). It "
        "measures the buffer available to absorb losses. Higher is safer.",
    "Cost-to-Income Ratio": "Operating expenses as a share of operating income — an efficiency "
        "measure. Lower means a leaner, more efficient bank.",
    "CAGR": "Compound Annual Growth Rate — the smoothed annual growth rate that would take a "
        "figure from its starting value to its ending value over the period.",
}

st.set_page_config(
    page_title="Indian Private Banks | Performance Analytics",
    page_icon="🏦",
    layout="wide",
    initial_sidebar_state="expanded",
)


# --------------------------------------------------------------------------- #
# Styling
# --------------------------------------------------------------------------- #
def inject_css() -> None:
    """Inject the Mountain Path Academy navy-and-gold theme."""
    st.markdown(
        """
        <style>
        :root { --navy:#0B2545; --blue:#123B65; --gold:#D4A017; --paper:#F6F8FC; }
        .stApp { background: linear-gradient(180deg,#F7F9FC 0%,#EEF3F8 100%); }
        [data-testid="stSidebar"] { background: linear-gradient(180deg,#0B2545,#153F69); }
        [data-testid="stSidebar"] * { color:#FFFFFF !important; }
        [data-testid="stSidebar"] .stMultiSelect span,
        [data-testid="stSidebar"] [data-baseweb="select"] * { color:#102A43 !important; }
        .hero { padding:1.65rem 2rem; border-radius:20px; color:white;
          background:linear-gradient(115deg,#081F3A 0%,#124A78 70%,#A97908 150%);
          box-shadow:0 12px 32px rgba(11,37,69,.18); margin-bottom:1rem; }
        .hero h1 { margin:0; font-size:2.05rem; line-height:1.2; color:white; }
        .hero p { margin:.55rem 0 0; color:#E7F0FA; font-size:1rem; }
        .eyebrow { color:#F3C84B; text-transform:uppercase; letter-spacing:.12em;
          font-weight:800; font-size:.75rem; margin-bottom:.55rem; }
        div[data-testid="stMetric"] { background:white; border:1px solid #DCE5EF;
          border-radius:14px; padding:1rem 1rem .8rem; box-shadow:0 6px 18px rgba(20,50,80,.06); }
        div[data-testid="stMetric"] label { color:#52677D !important; font-weight:700 !important; }
        .stTabs [data-baseweb="tab-list"] { gap:.45rem; background:#DDE7F1; padding:.4rem;
          border-radius:13px; overflow-x:auto; }
        .stTabs [data-baseweb="tab"] { background:#FFFFFF !important; border:1px solid #C8D6E5;
          border-radius:9px !important; color:#102A43 !important; font-weight:750; padding:.55rem 1rem; }
        .stTabs [aria-selected="true"] { background:#0B3B67 !important; color:#FFFFFF !important;
          border-color:#0B3B67 !important; }
        .stButton button, .stDownloadButton button { background:#0B3B67 !important;
          color:white !important; border:1px solid #0B3B67 !important; border-radius:9px !important;
          font-weight:750 !important; }
        .stButton button:hover, .stDownloadButton button:hover { background:#D4A017 !important;
          color:#071A2F !important; border-color:#D4A017 !important; }
        .section-title { font-size:1.3rem; font-weight:850; color:#0B2545; margin:.4rem 0 .65rem; }
        .subtle { color:#5A6B7D; font-size:.9rem; margin:-.35rem 0 .8rem; }
        .note { background:#FFF7DD; border-left:5px solid #D4A017; padding:.8rem 1rem;
          border-radius:9px; color:#4A3A0A; }
        .insight { background:#FFFFFF; border:1px solid #DCE5EF; border-left:5px solid #0B3B67;
          border-radius:11px; padding:.9rem 1.15rem; margin:.55rem 0; box-shadow:0 5px 15px rgba(20,50,80,.05); }
        .insight b { color:#0B2545; }
        .qtr-band { background:linear-gradient(120deg,#7A5A0B 0%,#B98A15 55%,#0B3B67 150%);
          color:#FFF8E6; border-radius:16px; padding:1rem 1.4rem; margin:.2rem 0 1rem;
          box-shadow:0 10px 26px rgba(122,90,11,.22); }
        .qtr-band .tag { display:inline-block; background:#FFFFFF; color:#7A5A0B; font-weight:850;
          font-size:.7rem; letter-spacing:.09em; text-transform:uppercase; padding:.2rem .6rem;
          border-radius:20px; margin-bottom:.4rem; }
        .qtr-band h2 { margin:.1rem 0 .15rem; font-size:1.4rem; color:#FFFFFF; }
        .qtr-band p { margin:0; color:#FBEFCB; font-size:.92rem; }
        /* Cleaner, high-contrast selectboxes (white field, navy text, gold focus) */
        div[data-testid="stSelectbox"] { margin:.35rem 0 1rem; }
        div[data-testid="stSelectbox"] label p {
          color:#0B2545 !important; font-size:1rem !important; font-weight:800 !important; }
        div[data-testid="stSelectbox"] [data-baseweb="select"] > div {
          min-height:48px !important; background:#FFFFFF !important;
          border:2px solid #0B3B67 !important; border-radius:10px !important;
          box-shadow:0 4px 12px rgba(11,37,69,.10) !important; }
        div[data-testid="stSelectbox"] [data-baseweb="select"] * {
          color:#0B2545 !important; -webkit-text-fill-color:#0B2545 !important; font-weight:700 !important; }
        div[data-testid="stSelectbox"] [data-baseweb="select"] > div:focus-within {
          border-color:#D4A017 !important; box-shadow:0 0 0 3px rgba(212,160,23,.28) !important; }
        [data-baseweb="popover"] [role="listbox"] { background:#FFFFFF !important;
          border:2px solid #0B3B67 !important; }
        [data-baseweb="popover"] [role="option"] { color:#0B2545 !important;
          -webkit-text-fill-color:#0B2545 !important; background:#FFFFFF !important; font-weight:600 !important; }
        [data-baseweb="popover"] [role="option"]:hover,
        [data-baseweb="popover"] [aria-selected="true"] { background:#0B3B67 !important;
          color:#FFFFFF !important; -webkit-text-fill-color:#FFFFFF !important; }
        section[data-testid="stSidebar"] div[data-testid="stSelectbox"] label p {
          color:#F3C84B !important; -webkit-text-fill-color:#F3C84B !important; }
        section[data-testid="stSidebar"] div[data-testid="stSelectbox"] [data-baseweb="select"] > div {
          background:#FFFFFF !important; border-color:#F3C84B !important; }
        .profile-card { background:linear-gradient(135deg,#071A2F,#123B65);
          border:1px solid rgba(243,200,75,.42); border-radius:14px; padding:17px;
          margin:15px 0 8px; box-shadow:0 7px 20px rgba(0,0,0,.18); }
        .profile-card .name { color:#F3C84B !important; font-weight:850; font-size:1rem; margin:0 0 5px; }
        .profile-card .title { color:#D7E9FA !important; font-size:.81rem; line-height:1.4; margin:0 0 8px; }
        .profile-card .stats { color:#AFC7DE !important; font-size:.76rem; line-height:1.45; margin:4px 0; }
        .profile-card .links { margin-top:11px; display:flex; gap:11px; flex-wrap:wrap; }
        .profile-card .links a { color:#F3C84B !important; text-decoration:none; font-size:.78rem; font-weight:750; }
        .profile-card .links a:hover { color:#FFFFFF !important; text-decoration:underline; }
        .about-section { background:linear-gradient(125deg,#0B2545,#123F69); color:#EAF3FC;
          border:1px solid rgba(212,160,23,.45); border-radius:17px; padding:26px 30px;
          margin:24px 0 12px; box-shadow:0 10px 27px rgba(11,37,69,.16); }
        .about-section h3 { color:#F3C84B !important; margin:0 0 11px; }
        .about-section p { color:#EAF3FC; line-height:1.62; margin:7px 0; }
        .about-section .highlight { color:#F3C84B; font-weight:800; }
        .academy-link { display:inline-block; margin-top:13px; padding:8px 16px;
          background:#D4A017; color:#071A2F !important; border-radius:8px;
          text-decoration:none; font-weight:850; }
        .mp-footer { text-align:center; padding:23px 0 8px; margin-top:25px;
          border-top:1px solid rgba(212,160,23,.4); color:#64778B; font-size:.84rem; }
        .mp-footer .footer-brand { color:#0B2545; font-size:1.12rem; font-weight:850; }
        .mp-footer .footer-profile { color:#38556F; margin:5px 0 8px; }
        .mp-footer a { color:#0B4F86; text-decoration:none; font-weight:750; margin:0 7px; }
        .mp-footer a:hover { color:#A97908; text-decoration:underline; }
        </style>
        """,
        unsafe_allow_html=True,
    )


# --------------------------------------------------------------------------- #
# Data loading and analytics helpers
# --------------------------------------------------------------------------- #
@st.cache_data
def load_data() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Load and type-clean the annual, quarterly and source datasets."""
    annual = pd.read_csv(DATA / "annual_bank_performance.csv", parse_dates=["period_end"])
    quarter = pd.read_csv(DATA / "q1_fy27.csv", parse_dates=["period_end"])
    sources = pd.read_csv(DATA / "sources.csv")
    for frame in (annual, quarter):
        numeric = frame.columns.difference(["bank", "period", "period_end", "status"])
        frame[numeric] = frame[numeric].apply(pd.to_numeric, errors="coerce")
    return annual, quarter, sources


def chart_layout(fig: go.Figure, height: int = 430) -> go.Figure:
    """Apply the shared clean chart styling."""
    fig.update_layout(
        height=height, margin=dict(l=20, r=20, t=55, b=20),
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="white",
        font=dict(family="Arial", color="#233B53"),
        legend_title_text="", hoverlabel=dict(bgcolor="white"),
    )
    fig.update_xaxes(showgrid=False)
    fig.update_yaxes(gridcolor="#E7EDF4", zeroline=False)
    return fig


def bar_by_bank(frame: pd.DataFrame, x: str, y: str, title: str,
                text_template: str = "%{text:,.2f}", labels: dict | None = None,
                height: int = 430) -> go.Figure:
    """Standardised per-bank bar chart with outside data labels."""
    fig = px.bar(frame, x=x, y=y, color="bank", text=y,
                 color_discrete_map=BANK_COLORS, title=title, labels=labels or {})
    fig.update_traces(texttemplate=text_template, textposition="outside", cliponaxis=False)
    return chart_layout(fig, height)


def line_by_bank(frame: pd.DataFrame, x: str, y: str, title: str,
                 labels: dict | None = None, height: int = 430) -> go.Figure:
    """Standardised per-bank line chart."""
    fig = px.line(frame, x=x, y=y, color="bank", markers=True,
                  color_discrete_map=BANK_COLORS, title=title, labels=labels or {})
    return chart_layout(fig, height)


def ranking(frame: pd.DataFrame, column: str, higher_better: bool) -> pd.DataFrame:
    """Return banks ranked by a metric (rank 1 = best)."""
    out = frame[["bank", column]].dropna().copy()
    if out.empty:
        return out.assign(Rank=pd.Series(dtype=int))
    out["Rank"] = out[column].rank(ascending=not higher_better, method="min").astype(int)
    return out.sort_values("Rank")


def leader(frame: pd.DataFrame, column: str, higher_better: bool) -> pd.Series | None:
    """Safely return the top-ranked row for a metric, or None if unavailable."""
    ranked = ranking(frame, column, higher_better)
    return None if ranked.empty else ranked.iloc[0]


def normalize(series: pd.Series, higher_better: bool) -> pd.Series:
    """Min-max normalise a series to 0-100, direction-aware. Flat series -> 100."""
    lo, hi = series.min(), series.max()
    if pd.isna(lo) or pd.isna(hi) or hi == lo:
        return pd.Series(100.0, index=series.index)
    scaled = (series - lo) / (hi - lo) * 100
    return scaled if higher_better else 100 - scaled


def composite_scores(latest: pd.DataFrame) -> pd.DataFrame:
    """Blend direction-aware normalised scores into a 0-100 composite per bank."""
    scored = latest[["bank"]].copy()
    contributions = []
    for _, col, higher in SCORE_METRICS:
        if col in latest and latest[col].notna().any():
            contributions.append(normalize(latest[col], higher))
    if not contributions:
        scored["Composite"] = np.nan
        return scored
    scored["Composite"] = pd.concat(contributions, axis=1).mean(axis=1).round(1)
    scored = scored.dropna(subset=["Composite"]).sort_values("Composite", ascending=False)
    scored["Rank"] = range(1, len(scored) + 1)
    return scored


def radar_figure(latest: pd.DataFrame, banks: list[str]) -> go.Figure:
    """Normalised multi-metric radar comparing the selected banks."""
    norm = {label: normalize(latest.set_index("bank")[col], higher)
            for label, col, higher in RADAR_METRICS if col in latest}
    dims = list(norm)
    fig = go.Figure()
    for bank in banks:
        values = [float(norm[d].get(bank, np.nan)) for d in dims]
        fig.add_trace(go.Scatterpolar(
            r=values + values[:1], theta=dims + dims[:1], fill="toself", name=bank,
            line=dict(color=BANK_COLORS.get(bank)), opacity=0.75))
    fig.update_layout(
        height=470, margin=dict(l=40, r=40, t=55, b=40),
        polar=dict(radialaxis=dict(range=[0, 100], gridcolor="#E1E8F1"),
                   angularaxis=dict(gridcolor="#E1E8F1")),
        paper_bgcolor="rgba(0,0,0,0)", font=dict(family="Arial", color="#233B53"),
        title="Peer radar — normalised score per dimension (100 = best in peer set)")
    return fig


def cagr_table(annual: pd.DataFrame) -> pd.DataFrame:
    """Compound annual growth rate for headline metrics, first to last available year."""
    rows = []
    for bank, grp in annual.sort_values("period_end").groupby("bank"):
        row = {"Bank": bank}
        for label, col in CAGR_METRICS.items():
            valid = grp.dropna(subset=[col])
            if len(valid) >= 2:
                start, end = valid[col].iloc[0], valid[col].iloc[-1]
                years = max((valid["period_end"].iloc[-1] - valid["period_end"].iloc[0]).days / 365.25, 1)
                row[label] = (end / start) ** (1 / years) - 1 if start > 0 else np.nan
            else:
                row[label] = np.nan
        rows.append(row)
    out = pd.DataFrame(rows)
    for label in CAGR_METRICS:
        out[label] = (out[label] * 100).round(1)
    return out


def correlation_figure(frame: pd.DataFrame) -> go.Figure:
    """Correlation heatmap across the metric panel (teaches how ratios move together)."""
    cols = ["net_profit_cr", "nii_cr", "nim_pct", "casa_pct", "gnpa_pct",
            "nnpa_pct", "roa_pct", "roe_pct", "car_pct", "cost_income_pct"]
    labels = ["Net profit", "NII", "NIM", "CASA", "GNPA", "NNPA",
              "ROA", "ROE", "CAR", "Cost/Income"]
    corr = frame[cols].corr().round(2)
    fig = px.imshow(corr, x=labels, y=labels, color_continuous_scale="RdBu", zmin=-1, zmax=1,
                    text_auto=True, title="How the metrics move together (Pearson correlation)")
    fig.update_layout(height=520, margin=dict(l=20, r=20, t=55, b=20),
                      paper_bgcolor="rgba(0,0,0,0)", font=dict(family="Arial", color="#233B53"))
    return fig


def trend_word(series: pd.Series) -> str:
    """Describe the direction of a short series in plain language."""
    valid = series.dropna()
    if len(valid) < 2:
        return "broadly stable"
    change = valid.iloc[-1] - valid.iloc[0]
    if abs(change) < 1e-9:
        return "broadly stable"
    return "rising" if change > 0 else "easing lower"


def commentary(bank: str, annual: pd.DataFrame, period: str) -> str:
    """Generate a short template-based narrative for one bank at a given period."""
    grp = annual[annual["bank"] == bank].sort_values("period_end")
    if grp.empty:
        return f"No data available for {bank}."
    latest = grp[grp["period"] == period]
    if latest.empty:
        latest = grp.tail(1)
    row = latest.iloc[0]
    prior = grp[grp["period_end"] < row["period_end"]].tail(1)

    parts = [f"<b>{bank}</b> reported a net profit of ₹{row['net_profit_cr']:,.0f} crore in {row['period']}"]
    if not prior.empty and prior["net_profit_cr"].notna().all() and prior["net_profit_cr"].iloc[0] > 0:
        yoy = (row["net_profit_cr"] / prior["net_profit_cr"].iloc[0] - 1) * 100
        parts[0] += f", a year-on-year change of {yoy:+.1f}%"
    parts[0] += "."

    parts.append(
        f"Its net interest margin stood at {row['nim_pct']:.2f}% with a CASA ratio of "
        f"{row['casa_pct']:.1f}%, while profitability was reflected in an ROA of "
        f"{row['roa_pct']:.2f}% and an ROE of {row['roe_pct']:.1f}%.")

    gnpa_dir = trend_word(grp["gnpa_pct"])
    parts.append(
        f"Asset quality shows gross NPAs at {row['gnpa_pct']:.2f}% (net {row['nnpa_pct']:.2f}%), "
        f"{gnpa_dir} across the period, on a capital adequacy of {row['car_pct']:.1f}% and a "
        f"cost-to-income ratio of {row['cost_income_pct']:.1f}%.")
    return " ".join(parts)


def bank_profiles_markdown(annual: pd.DataFrame, banks: list[str], period: str) -> str:
    """Build a downloadable one-page markdown profile for the selected banks."""
    lines = [f"# Indian Private Banks — One-Page Profiles ({period})",
             "_Prepared with the Mountain Path Academy Banking Analytics dashboard._", ""]
    for bank in banks:
        grp = annual[annual["bank"] == bank]
        latest = grp[grp["period"] == period]
        if latest.empty:
            continue
        r = latest.iloc[0]
        lines += [
            f"## {bank}", "",
            f"- Net profit: ₹{r['net_profit_cr']:,.0f} crore",
            f"- Net interest income: ₹{r['nii_cr']:,.0f} crore  |  NIM: {r['nim_pct']:.2f}%",
            f"- Advances: ₹{r['advances_cr']:,.0f} crore  |  Deposits: ₹{r['deposits_cr']:,.0f} crore",
            f"- CASA: {r['casa_pct']:.1f}%  |  Cost-to-income: {r['cost_income_pct']:.1f}%",
            f"- Gross NPA: {r['gnpa_pct']:.2f}%  |  Net NPA: {r['nnpa_pct']:.2f}%",
            f"- ROA: {r['roa_pct']:.2f}%  |  ROE: {r['roe_pct']:.1f}%  |  CAR: {r['car_pct']:.1f}%",
            "", f"> {commentary(bank, annual, period).replace('<b>', '**').replace('</b>', '**')}", "",
        ]
    lines.append("\n_Ratios are educational and do not constitute investment advice._")
    return "\n".join(lines)


def workbook_bytes(annual: pd.DataFrame, quarter: pd.DataFrame, sources: pd.DataFrame) -> bytes:
    """Export the datasets to a formatted multi-sheet Excel workbook."""
    output = BytesIO()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0B3B67")
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        annual.to_excel(writer, index=False, sheet_name="Annual FY22-FY26")
        quarter.to_excel(writer, index=False, sheet_name="Q1 FY2027")
        sources.to_excel(writer, index=False, sheet_name="Sources")
        for ws in writer.book.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            ws.sheet_view.showGridLines = False
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
            for col in ws.columns:
                letter = col[0].column_letter
                width = min(max(len(str(c.value or "")) for c in col) + 2, 42)
                ws.column_dimensions[letter].width = width
    return output.getvalue()


@st.cache_data(ttl=1800, show_spinner=False)
def market_prices(tickers: tuple[str, ...], start: str, end: str) -> pd.DataFrame:
    """Fetch adjusted close prices from Yahoo Finance; return empty frame on any failure."""
    try:
        import yfinance as yf
        raw = yf.download(list(tickers), start=start, end=end, auto_adjust=True, progress=False)
        if raw.empty:
            return pd.DataFrame()
        close = raw["Close"] if isinstance(raw.columns, pd.MultiIndex) else raw[["Close"]]
        if isinstance(close, pd.Series):
            close = close.to_frame(name=tickers[0])
        return close.dropna(how="all")
    except Exception:
        return pd.DataFrame()


def kpi(container, label: str, bank_row: pd.Series | None, value_fmt: str,
        delta: str | None = None, delta_color: str = "normal") -> None:
    """Render a KPI metric card with a graceful fallback when data is missing."""
    if bank_row is None:
        container.metric(label, "—", "no data")
        return
    container.metric(label, bank_row["bank"], delta if delta else value_fmt,
                     delta_color=delta_color)


def yoy_delta(frame: pd.DataFrame, bank: str, col: str, period: str, fmt: str) -> str | None:
    """Compute a formatted year-on-year delta string for a bank/metric, if possible."""
    grp = frame[frame["bank"] == bank].sort_values("period_end")
    cur = grp[grp["period"] == period]
    if cur.empty:
        return None
    prior = grp[grp["period_end"] < cur["period_end"].iloc[0]].tail(1)
    if prior.empty or pd.isna(prior[col].iloc[0]) or pd.isna(cur[col].iloc[0]):
        return None
    return fmt.format(cur[col].iloc[0] - prior[col].iloc[0])


# --------------------------------------------------------------------------- #
# Application layout
# --------------------------------------------------------------------------- #
inject_css()
annual, quarter, sources = load_data()

st.markdown(
    """<div class="hero"><div class="eyebrow">The Mountain Path Academy · Banking Analytics</div>
    <h1>Performance Analysis of India’s Leading Private Banks</h1>
    <p>ICICI Bank · HDFC Bank · Axis Bank · Kotak Mahindra Bank · Yes Bank</p></div>""",
    unsafe_allow_html=True,
)

all_banks = list(BANK_COLORS)
with st.sidebar:
    st.markdown("## Analysis controls")
    selected = st.multiselect("Banks", all_banks, default=all_banks)
    if not selected:
        st.warning("Select at least one bank — showing all.")
        selected = all_banks
    period = st.selectbox("Latest annual snapshot", annual["period"].drop_duplicates().tolist()[::-1])
    st.markdown("---")
    st.caption("Figures in ₹ crore unless otherwise stated. Review the Sources tab before citing.")
    st.markdown(
        """<div class='profile-card'>
        <p class='name'>Prof. V. Ravichandran</p>
        <p class='title'>Visiting Professor &amp; Professor of Practice at Leading Business Schools<br>
        Founder — The Mountain Path Academy</p>
        <p class='stats'>28+ yrs · HSBC Global Banking &amp; Markets · Synechron<br>
        12+ yrs teaching · Financial Risk · Derivatives · ALM</p>
        <div class='links'>
          <a href='https://themountainpathacademy.com' target='_blank'>🏔️ Academy</a>
          <a href='https://www.linkedin.com/in/trichyravis' target='_blank'>💼 LinkedIn</a>
          <a href='https://github.com/trichyravis' target='_blank'>💻 GitHub</a>
        </div></div>""",
        unsafe_allow_html=True,
    )

a = annual[annual["bank"].isin(selected)].copy()
q = quarter[quarter["bank"].isin(selected)].copy()
latest = a[a["period"] == period]

tabs = st.tabs([
    "Executive Dashboard", "★ Q1 FY2027 · Apr–Jun 2026", "Annual Trends", "NIM, NII & CASA Growth",
    "Asset Quality & Correlations", "Market Performance", "Insights & Glossary", "Downloads & Sources",
])

# ----------------------------- Executive Dashboard ------------------------- #
with tabs[0]:
    # Latest-quarter spotlight — the freshest data, featured first.
    st.markdown(
        '<div class="qtr-band"><span class="tag">Latest quarter · most recent results</span>'
        '<h2>Q1 FY2027 — Apr–Jun 2026</h2>'
        '<p>The newest reported quarter (ended 30 June 2026). Full detail in the '
        '“Q1 FY2027 · Apr–Jun 2026” tab.</p></div>',
        unsafe_allow_html=True,
    )
    if not q.empty:
        qc1, qc2, qc3, qc4 = st.columns(4)
        _qp = leader(q, "net_profit_cr", True)
        _qg = leader(q, "advances_growth_yoy_pct", True)
        _qn = leader(q, "nim_pct", True)
        _qq = leader(q, "gnpa_pct", False)
        if _qp is not None: qc1.metric("Q1 profit leader", _qp["bank"], f"₹{_qp['net_profit_cr']:,.0f} Cr")
        if _qg is not None: qc2.metric("Q1 advances growth", _qg["bank"], f"{_qg['advances_growth_yoy_pct']:.1f}% YoY")
        if _qn is not None: qc3.metric("Q1 NIM leader", _qn["bank"], f"{_qn['nim_pct']:.2f}%")
        if _qq is not None: qc4.metric("Q1 lowest GNPA", _qq["bank"], f"{_qq['gnpa_pct']:.2f}%")

    st.markdown('<div class="section-title">Executive performance snapshot</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="subtle">Leaders for {period}, with the leader’s year-on-year movement.</div>',
                unsafe_allow_html=True)
    if latest.empty:
        st.info("No annual data for the chosen banks and period.")
    else:
        c1, c2, c3, c4 = st.columns(4)
        pl, nl, ql, rl = (leader(latest, "net_profit_cr", True), leader(latest, "nim_pct", True),
                          leader(latest, "gnpa_pct", False), leader(latest, "roa_pct", True))
        kpi(c1, "Highest net profit", pl, f"₹{pl['net_profit_cr']:,.0f} Cr" if pl is not None else "",
            yoy_delta(a, pl["bank"], "net_profit_cr", period, "{:+,.0f} Cr YoY") if pl is not None else None)
        kpi(c2, "Highest NIM", nl, f"{nl['nim_pct']:.2f}%" if nl is not None else "",
            yoy_delta(a, nl["bank"], "nim_pct", period, "{:+.2f} pp YoY") if nl is not None else None)
        kpi(c3, "Lowest gross NPA", ql, f"{ql['gnpa_pct']:.2f}%" if ql is not None else "",
            yoy_delta(a, ql["bank"], "gnpa_pct", period, "{:+.2f} pp YoY") if ql is not None else None,
            delta_color="inverse")
        kpi(c4, "Highest ROA", rl, f"{rl['roa_pct']:.2f}%" if rl is not None else "",
            yoy_delta(a, rl["bank"], "roa_pct", period, "{:+.2f} pp YoY") if rl is not None else None)

        st.markdown('<div class="section-title">Composite performance scorecard</div>', unsafe_allow_html=True)
        st.markdown('<div class="subtle">Each bank scored 0–100 on eight direction-aware metrics '
                    '(profitability, margin, funding, asset quality, capital, efficiency), then averaged.</div>',
                    unsafe_allow_html=True)
        scores = composite_scores(latest)
        sc1, sc2 = st.columns([1.25, 1])
        if not scores.empty:
            fig = px.bar(scores, x="Composite", y="bank", orientation="h", color="bank",
                         color_discrete_map=BANK_COLORS, text="Composite",
                         title="Composite score (higher is stronger)")
            fig.update_traces(texttemplate="%{text:.1f}", textposition="outside", cliponaxis=False)
            fig.update_layout(yaxis=dict(categoryorder="total ascending"))
            sc1.plotly_chart(chart_layout(fig), width="stretch")
            sc2.dataframe(scores[["Rank", "bank", "Composite"]].rename(columns={"bank": "Bank"}),
                          hide_index=True, width="stretch")
        if len(selected) >= 2:
            st.plotly_chart(radar_figure(latest, selected), width="stretch")

        st.markdown('<div class="section-title">Compare a single parameter</div>', unsafe_allow_html=True)
        metric_label = st.selectbox("Compare parameter", list(METRICS), index=0, key="exec_metric")
        col, higher, _ = METRICS[metric_label]
        rank = ranking(latest, col, higher)
        if not rank.empty:
            st.plotly_chart(bar_by_bank(rank, "bank", col, f"{metric_label} · {period}"), width="stretch")
            st.dataframe(rank.rename(columns={"bank": "Bank", col: metric_label}),
                         hide_index=True, width="stretch")

# ----------------------------- Annual Trends ------------------------------- #
with tabs[2]:
    st.markdown('<div class="section-title">Five-year audited trend analysis</div>', unsafe_allow_html=True)
    metric_label = st.selectbox("Trend parameter", list(METRICS), index=0, key="trend_metric")
    col, _, _ = METRICS[metric_label]
    trend = a.dropna(subset=[col])
    if not trend.empty:
        st.plotly_chart(line_by_bank(trend, "period", col, f"{metric_label}: FY2022–FY2026"),
                        width="stretch")
        pivot = trend.pivot(index="bank", columns="period", values=col).reset_index()
        st.dataframe(pivot.rename(columns={"bank": "Bank"}), hide_index=True, width="stretch")

    st.markdown('<div class="section-title">Five-year growth (CAGR)</div>', unsafe_allow_html=True)
    st.markdown('<div class="subtle">Compound annual growth rate from the first to the latest audited year.</div>',
                unsafe_allow_html=True)
    cagr = cagr_table(a)
    if not cagr.empty:
        long = cagr.melt(id_vars="Bank", var_name="Metric", value_name="CAGR %")
        fig = px.bar(long, x="Metric", y="CAGR %", color="Bank", barmode="group",
                     color_discrete_map=BANK_COLORS, text="CAGR %", title="Headline CAGR by bank")
        fig.update_traces(texttemplate="%{text:.1f}%", textposition="outside", cliponaxis=False)
        st.plotly_chart(chart_layout(fig), width="stretch")
        st.dataframe(cagr, hide_index=True, width="stretch")

# ----------------------------- NIM/NII/CASA Growth ------------------------- #
with tabs[3]:
    st.markdown('<div class="section-title">NIM, NII and CASA growth analysis</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="note"><b>Interpretation:</b> NII growth is measured in percentage terms. '
        'NIM and CASA are ratios, so their change is shown in basis points (100 bps = 1 percentage point). '
        'A higher NIM is generally positive; CASA must be read with deposit growth and funding cost.</div>',
        unsafe_allow_html=True,
    )
    growth = a.sort_values(["bank", "period"]).copy()
    growth["nii_growth_yoy_pct"] = growth.groupby("bank")["nii_cr"].pct_change(fill_method=None) * 100
    growth["nim_change_yoy_bps"] = growth.groupby("bank")["nim_pct"].diff() * 100
    growth["casa_change_yoy_bps"] = growth.groupby("bank")["casa_pct"].diff() * 100
    latest_growth = growth[growth["period"] == period]

    g1, g2, g3 = st.columns(3)
    if latest_growth["nii_growth_yoy_pct"].notna().any():
        lead = latest_growth.loc[latest_growth["nii_growth_yoy_pct"].idxmax()]
        g1.metric("Fastest NII growth", lead["bank"], f"{lead['nii_growth_yoy_pct']:.1f}% YoY")
    if latest_growth["nim_change_yoy_bps"].notna().any():
        lead = latest_growth.loc[latest_growth["nim_change_yoy_bps"].idxmax()]
        g2.metric("Best NIM movement", lead["bank"], f"{lead['nim_change_yoy_bps']:+.0f} bps")
    if latest_growth["casa_change_yoy_bps"].notna().any():
        lead = latest_growth.loc[latest_growth["casa_change_yoy_bps"].idxmax()]
        g3.metric("Best CASA movement", lead["bank"], f"{lead['casa_change_yoy_bps']:+.0f} bps")

    analysis_choice = st.radio(
        "Select analysis", ["NII growth", "NIM level and movement", "CASA level and movement"],
        horizontal=True,
    )
    if analysis_choice == "NII growth":
        left, right = st.columns([1.35, 1])
        left.plotly_chart(line_by_bank(growth, "period", "nii_cr", "Net interest income trend",
                          labels={"nii_cr": "NII (₹ crore)", "period": "Financial year"}), width="stretch")
        right.plotly_chart(bar_by_bank(latest_growth.dropna(subset=["nii_growth_yoy_pct"]), "bank",
                           "nii_growth_yoy_pct", f"NII growth · {period}", "%{text:+.1f}%",
                           {"nii_growth_yoy_pct": "YoY growth (%)", "bank": "Bank"}), width="stretch")
    else:
        is_nim = analysis_choice.startswith("NIM")
        level_col = "nim_pct" if is_nim else "casa_pct"
        change_col = "nim_change_yoy_bps" if is_nim else "casa_change_yoy_bps"
        title = "Net interest margin" if is_nim else "CASA ratio"
        left, right = st.columns([1.35, 1])
        left.plotly_chart(line_by_bank(growth, "period", level_col, f"{title} trend",
                          labels={level_col: f"{title} (%)", "period": "Financial year"}), width="stretch")
        right.plotly_chart(bar_by_bank(latest_growth.dropna(subset=[change_col]), "bank", change_col,
                           f"{title} movement · {period}", "%{text:+.0f} bps",
                           {change_col: "YoY change (bps)", "bank": "Bank"}), width="stretch")

    growth_table = latest_growth[["bank", "nii_cr", "nii_growth_yoy_pct", "nim_pct",
                                  "nim_change_yoy_bps", "casa_pct", "casa_change_yoy_bps"]].copy()
    growth_table.columns = ["Bank", "NII (₹ Cr)", "NII Growth YoY (%)", "NIM (%)",
                            "NIM Change (bps)", "CASA (%)", "CASA Change (bps)"]
    st.dataframe(growth_table.round(2), hide_index=True, width="stretch")
    if "HDFC Bank" in selected and period in {"FY2024", "FY2025"}:
        st.caption("Interpret the HDFC Bank trend in light of the HDFC Ltd merger and the enlarged "
                   "post-merger balance sheet.")

# ----------------------------- Q1 FY2027 (latest quarter) ------------------ #
with tabs[1]:
    st.markdown(
        '<div class="qtr-band"><span class="tag">Most recent results</span>'
        '<h2>Q1 FY2027 — Apr–Jun 2026</h2>'
        '<p>Quarter ended 30 June 2026 — the freshest reported numbers in this dashboard.</p></div>',
        unsafe_allow_html=True,
    )
    st.markdown('<div class="note">Quarterly results are unaudited/limited-review figures. Annual and '
                'quarterly values are intentionally presented in separate tabs.</div>', unsafe_allow_html=True)
    st.write("")
    if q.empty:
        st.info("No quarterly data for the chosen banks.")
    else:
        c1, c2, c3, c4 = st.columns(4)
        qp, qg = leader(q, "net_profit_cr", True), leader(q, "advances_growth_yoy_pct", True)
        qn, qgn = leader(q, "nim_pct", True), leader(q, "gnpa_pct", False)
        if qp is not None: c1.metric("Profit leader", qp["bank"], f"₹{qp['net_profit_cr']:,.0f} Cr")
        if qg is not None: c2.metric("Advances growth leader", qg["bank"], f"{qg['advances_growth_yoy_pct']:.1f}% YoY")
        if qn is not None: c3.metric("NIM leader", qn["bank"], f"{qn['nim_pct']:.2f}%")
        if qgn is not None: c4.metric("Lowest GNPA", qgn["bank"], f"{qgn['gnpa_pct']:.2f}%")
        q_options = {
            "Net Profit (₹ crore)": ("net_profit_cr", True), "NII (₹ crore)": ("nii_cr", True),
            "NII Growth YoY (%)": ("nii_growth_yoy_pct", True), "NIM (%)": ("nim_pct", True),
            "Advances Growth YoY (%)": ("advances_growth_yoy_pct", True),
            "Deposits Growth YoY (%)": ("deposits_growth_yoy_pct", True), "Gross NPA (%)": ("gnpa_pct", False),
            "Net NPA (%)": ("nnpa_pct", False), "ROA (%)": ("roa_pct", True), "ROE (%)": ("roe_pct", True),
            "Provisions (₹ crore)": ("provisions_cr", False),
        }
        q_label = st.selectbox("Quarterly parameter", q_options, key="q_metric")
        q_col, q_high = q_options[q_label]
        q_rank = ranking(q, q_col, q_high)
        if not q_rank.empty:
            st.plotly_chart(bar_by_bank(q_rank, "bank", q_col, f"{q_label} · Q1 FY2027"), width="stretch")
            st.dataframe(q_rank.rename(columns={"bank": "Bank", q_col: q_label}),
                         hide_index=True, width="stretch")

        st.markdown('<div class="section-title">Q1 balance-sheet momentum — advances vs deposits (YoY)</div>',
                    unsafe_allow_html=True)
        gcols = ["bank", "advances_growth_yoy_pct", "deposits_growth_yoy_pct"]
        if set(gcols).issubset(q.columns):
            gm = q[gcols].dropna(how="all", subset=gcols[1:]).melt(
                id_vars="bank", var_name="Measure", value_name="YoY growth (%)")
            gm["Measure"] = gm["Measure"].map({"advances_growth_yoy_pct": "Advances",
                                               "deposits_growth_yoy_pct": "Deposits"})
            fig = px.bar(gm, x="bank", y="YoY growth (%)", color="Measure", barmode="group",
                         text="YoY growth (%)", color_discrete_map={"Advances": "#0B3B67", "Deposits": "#D4A017"},
                         title="Advances vs deposits growth · Q1 FY2027 (Apr–Jun 2026)",
                         labels={"bank": "Bank"})
            fig.update_traces(texttemplate="%{text:.1f}%", textposition="outside", cliponaxis=False)
            st.plotly_chart(chart_layout(fig), width="stretch")

# ----------------------------- Asset Quality & Correlations ---------------- #
with tabs[4]:
    st.markdown('<div class="section-title">Asset quality and capital resilience</div>', unsafe_allow_html=True)
    left, right = st.columns(2)
    for container, col, title in [(left, "gnpa_pct", "Gross NPA (%)"), (right, "nnpa_pct", "Net NPA (%)")]:
        container.plotly_chart(line_by_bank(a, "period", col, title, height=390), width="stretch")
    if not latest.empty:
        fig = px.scatter(latest, x="car_pct", y="roa_pct", size="advances_cr", color="bank",
                         color_discrete_map=BANK_COLORS, hover_name="bank",
                         title=f"Capital adequacy versus return on assets · {period}",
                         labels={"car_pct": "Capital adequacy (%)", "roa_pct": "ROA (%)"})
        st.plotly_chart(chart_layout(fig), width="stretch")

    st.markdown('<div class="section-title">How the metrics relate</div>', unsafe_allow_html=True)
    st.markdown('<div class="subtle">Correlation across every bank-year in the selection. Deep blue = strong '
                'positive, deep red = strong negative. A teaching aid for how ratios interact.</div>',
                unsafe_allow_html=True)
    if len(a) >= 3:
        st.plotly_chart(correlation_figure(a), width="stretch")
    else:
        st.info("Select more banks to compute a meaningful correlation matrix.")

# ----------------------------- Market Performance -------------------------- #
with tabs[5]:
    st.markdown('<div class="section-title">NSE shareholder-return analysis</div>', unsafe_allow_html=True)
    m1, m2 = st.columns(2)
    start = m1.date_input("Start date", value=pd.Timestamp("2022-04-01"), max_value=pd.Timestamp.today())
    end = m2.date_input("End date", value=pd.Timestamp.today(), max_value=pd.Timestamp.today())
    if start >= end:
        st.error("Start date must be before end date.")
    else:
        ticker_list = tuple(TICKERS[b] for b in selected)
        prices = market_prices(ticker_list, str(start), str(pd.Timestamp(end) + pd.Timedelta(days=1)))
        if prices.empty:
            st.warning("Live market data is temporarily unavailable. The financial-analysis tabs remain "
                       "fully functional.")
        else:
            prices = prices.rename(columns={v: k for k, v in TICKERS.items()})
            normalized = prices.div(prices.iloc[0]).mul(100)
            long = normalized.reset_index().melt(id_vars=normalized.index.name or "Date",
                                                 var_name="Bank", value_name="Indexed value")
            xname = long.columns[0]
            fig = px.line(long, x=xname, y="Indexed value", color="Bank", color_discrete_map=BANK_COLORS,
                          title="Indexed share-price performance (start = 100)")
            st.plotly_chart(chart_layout(fig), width="stretch")
            returns = prices.pct_change().dropna()
            summary = pd.DataFrame({
                "Total return (%)": (prices.iloc[-1] / prices.iloc[0] - 1) * 100,
                "Annualised volatility (%)": returns.std() * np.sqrt(252) * 100,
                "Sharpe ratio (RF=0%)": returns.mean() / returns.std() * np.sqrt(252),
            }).replace([np.inf, -np.inf], np.nan).round(2)
            st.dataframe(summary.reset_index(names="Bank"), hide_index=True, width="stretch")
            st.caption("Sharpe ratio assumes a 0% risk-free rate for classroom comparability. Market data "
                       "may be delayed and is not investment advice.")

# ----------------------------- Insights & Glossary ------------------------- #
with tabs[6]:
    st.markdown('<div class="section-title">Automated commentary</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="subtle">Plain-language read-out for each selected bank ({period}).</div>',
                unsafe_allow_html=True)
    for bank in selected:
        st.markdown(f'<div class="insight">{commentary(bank, a, period)}</div>', unsafe_allow_html=True)

    st.markdown('<div class="section-title">Glossary of ratios</div>', unsafe_allow_html=True)
    for term, definition in GLOSSARY.items():
        with st.expander(term):
            st.write(definition)

# ----------------------------- Downloads & Sources ------------------------- #
with tabs[7]:
    st.markdown('<div class="section-title">Download data and review sources</div>', unsafe_allow_html=True)
    excel = workbook_bytes(a, q, sources)
    d1, d2, d3 = st.columns(3)
    d1.download_button("Download complete Excel", excel, "private_banks_performance.xlsx",
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", width="stretch")
    d2.download_button("Download annual CSV", a.to_csv(index=False).encode(), "annual_bank_performance.csv",
                       "text/csv", width="stretch")
    d3.download_button("Download Q1 FY2027 CSV", q.to_csv(index=False).encode(), "q1_fy27.csv",
                       "text/csv", width="stretch")
    st.download_button("Download one-page bank profiles (Markdown)",
                       bank_profiles_markdown(a, selected, period).encode(),
                       f"bank_profiles_{period}.md", "text/markdown")

    st.markdown("#### Data sources")
    st.dataframe(sources, hide_index=True, width="stretch", column_config={
        "source_url": st.column_config.LinkColumn("Source link", display_text="Open source")
    })
    st.info("Research note: verify transcribed figures against the linked filings before publication or "
            "investment use. Market data may be delayed and is not investment advice.")

# ----------------------------- About / footer ------------------------------ #
st.markdown(
    """<div class='about-section'>
    <h3>About This Project</h3>
    <p>Developed by <span class='highlight'>Prof. V. Ravichandran</span>, Visiting Professor &amp;
    Professor of Practice at Leading Business Schools and founder of
    <span class='highlight'>The Mountain Path Academy</span>.</p>
    <p>With <span class='highlight'>28+ years of industry experience</span> at HSBC Global Banking &amp;
    Markets and Synechron, and <span class='highlight'>12+ years of teaching</span> Financial Risk
    Management, Derivatives, Fixed Income Securities, Financial Modelling and ALM, this dashboard
    brings a practitioner-educator approach to Indian banking performance analysis.</p>
    <a class='academy-link' href='https://themountainpathacademy.com' target='_blank'>🏔️ Visit The Mountain Path Academy</a>
    </div>
    <div class='mp-footer'>
      <div class='footer-brand'>🏔️ The Mountain Path Academy</div>
      <div class='footer-profile'>Prof. V. Ravichandran · Visiting Professor &amp; Professor of Practice at Leading Business Schools</div>
      <div><a href='https://themountainpathacademy.com' target='_blank'>themountainpathacademy.com</a></div>
      <div style='margin-top:8px'>
        <a href='https://www.linkedin.com/in/trichyravis' target='_blank'>LinkedIn</a>
        <a href='https://github.com/trichyravis' target='_blank'>GitHub</a>
      </div>
      <div style='margin-top:10px;font-size:.77rem'>Educational analytics project · Ratios do not constitute investment advice · © 2026 The Mountain Path Academy</div>
    </div>""",
    unsafe_allow_html=True,
)
